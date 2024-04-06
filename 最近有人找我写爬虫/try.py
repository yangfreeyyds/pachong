import asyncio
import json
from time import sleep
import requests
import urllib.parse
from openpyxl import load_workbook
import random

def get_proxies():
    # 提取代理API接口，获取1个代理IP
    api_url = "https://dps.kdlapi.com/api/getdps/?secret_id=o2j3rj83e51tl6bc26sc&signature=bau1xi3u8f4w4c4dze96k1qogsbhyn6f&num=1&pt=1&sep=1"

    # 获取API接口返回的代理IP
    proxy_ip = requests.get(api_url).text

    # 用户名密码认证(私密代理/独享代理)
    username = "d3967661328"
    password = "154vkktm"
    proxies = {
        # "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip},
        "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip},
    }
    return proxies
user_agentpp = "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60"
user_agent = [
    # Opera
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60",
    "Opera/8.0 (Windows NT 5.1; U; en)",
    "Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50",

    # Firefox
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0",
    "Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10",

    # Safari
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2",

    # chrome
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16",

    # 360
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko",

    # 淘宝浏览器
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11",

    # 猎豹浏览器
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E; LBBROWSER)",

    # QQ浏览器
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; QQBrowser/7.0.3698.400)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E) ",

    # sogou浏览器
    "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0)",

    # maxthon浏览器
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Maxthon/4.4.3.4000 Chrome/30.0.1599.101 Safari/537.36",

    # UC浏览器
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/38.0.2125.122 UBrowser/4.0.3214.0 Safari/537.36",

    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.100 Safari/537.36",

    "Mozilla/5.0 (iPhone; CPU iPhone OS 13_3_1 like Mac OS X; zh-CN) AppleWebKit/537.51.1 (KHTML, like Gecko) Mobile/17D50 UCBrowser/12.8.2.1268 Mobile AliApp(TUnionSDK/0.1.20.3)",

    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.116 Safari/537.36",

    "Mozilla/5.0 (Linux; Android 8.1.0; OPPO R11t Build/OPM1.171019.011; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/76.0.3809.89 Mobile Safari/537.36 T7/11.19 SP-engine/2.15.0 baiduboxapp/11.19.5.10 (Baidu; P1 8.1.0)",

    "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36",

    "Mozilla/5.0 (iPhone; CPU iPhone OS 13_3_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 SP-engine/2.14.0 main%2F1.0 baiduboxapp/11.18.0.16 (Baidu; P2 13.3.1) NABar/0.0 ",

    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134",

    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36",

    "Mozilla/5.0 (iPhone; CPU iPhone OS 12_4_4 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 MicroMessenger/7.0.10(0x17000a21) NetType/4G Language/zh_CN",

    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36",

    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36",

    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36",

    "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.108 Safari/537.36",

    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36",

    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.106 Safari/537.36"

]
def get():
    user = {}
    user['User-Agent'] = user_agent[random.randint(0, len(user_agent) - 1)]
    return str(user)

#逆向xbogus,ttwid
def dataget():
    # 目标URL
    url = "https://tuu-yangfreeyyds-projects.vercel.app"

    # 请求头，指定Content-Type为application/json
    headers = {
        "Content-Type": "application/json"
    }
    user_agentpp = get()
    # 请求体数据
    data = {
        "url": "https://www.douyin.com/aweme/v1/web/aweme/detail/?aid=6383&version_name=19.6.0&device_platform=webapp",
        "userAgent": user_agentpp
    }

    # 发送POST请求
    response = requests.post(url, headers=headers, data=json.dumps(data))
    # 提取
    data = response.json()
    data = data.get("data")
    xbogus = data.get('xbogus')
    mstoken = data.get('mstoken')
    ttwid = data.get('ttwid')
    return xbogus,mstoken,ttwid
#暂时没用
def get_ms_token():
    """
    根据传入长度产生随机字符串
    """
    randomlength = 107
    random_str = ''
    base_str = 'ABCDEFGHIGKLMNOPQRSTUVWXYZabcdefghigklmnopqrstuvwxyz0123456789='
    length = len(base_str) - 1
    for _ in range(randomlength):
        random_str += base_str[random.randint(0, length)]
    return random_str

#print(get_ms_token())

def payapa(keyword,proxies):
    xbogus, mstoken, ttwid = dataget()
    encoded_keyword = urllib.parse.quote(keyword, encoding='utf-8')
    offset = 0
    sum = 0
    while offset < 29:
        url = "https://www.douyin.com/aweme/v1/web/general/search/single/"
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en-US,en;q=0.5',
            'Connection': 'keep-alive',
            'Cookie': 'xbogus='+xbogus+'; mstoken='+mstoken+'; ttwid='+ttwid+'; IsDouyinActive="true"; ',
            # 'Cookie': '__ac_nonce=0660fad8700bc0118153; __ac_signature=_02B4Z6wo00f01mvyBagAAIDDKMKcyhGJ8B5rwiEAAP0C94; SEARCH_RESULT_LIST_TYPE=%22single%22; ttwid=1%7C0niGoYF4S3w-c4NWj3rgCcc3762H08TSb8VU8u0mdXA%7C1712303495%7C94211a887138f2a34721350081eae806cdbdd263f85e1e306270d1c61003f3bc; home_can_add_dy_2_desktop=%220%22; IsDouyinActive=true; csrf_session_id=a652f0fad14330c0dccc7b46a3047ffe; FORCE_LOGIN=%7B%22videoConsumedRemainSeconds%22%3A180%7D; msToken=AMKz8pJMzmiaCSB8bv891ewgHONhyp07GiLcGkcGR9NmR73rtClC10HHcP7y5hctciTa4I77cwcUlD37BwnK57wuvEADaOk3S2CSUCVC5sn0R3lMlehBmtPiHdsXWQq3; volume_info=%7B%22isUserMute%22%3Afalse%2C%22isMute%22%3Afalse%2C%22volume%22%3A0.5%7D',
            'Host': 'www.douyin.com',
            'Referer': 'https://www.douyin.com/search/' + encoded_keyword + '?type=general',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'TE': 'trailers',
            'User-Agent': user_agentpp
        }
        params = {
            "device_platform": "webapp",
            "aid": "6383",
            "channel": "channel_pc_web",
            "search_channel": "aweme_general",
            "enable_history": "1",
            "filter_selected": '{"sort_type":"1","publish_time":"0"}',
            "keyword": keyword,
            "search_source": "tab_search",
            "query_correct_type": "1",
            "is_filter_search": "1",
            "from_group_id": "",
            "offset": offset,
            "count": "10",
        }
        dict = {
            'xbogus':xbogus,
            'mstoken':mstoken,
            'ttwid':ttwid,
        }
        response = requests.get(url, params=params, proxies=proxies, headers=headers, timeout=4, cookies=dict)
        print(response.status_code)
        data = response.json()
        if str(data.get('data',[])) == '[]' and offset != 0:
            offset += 10
            response.close()
            break
        elif str(data.get('data',[])) == '[]' and offset == 0:
            response.close()
            print(data)
            continue
        else:
            for item in data.get("data", []):
                try:
                    aweme_info = item.get("aweme_info", {})
                    # 注意：如果aweme_info不是字典，下一行将抛出异常
                    statistics = aweme_info.get("statistics", {})
                    # 确保 'digg_count' 存在并且是整数，否则使用默认值 0
                    digg_count = int(statistics.get('digg_count', 0))
                    sum += digg_count
                except ValueError:
                    # 如果digg_count不是一个可以转换为整数的值，打印错误并跳过
                    print("ValueError: digg_count could not be converted to an integer.")
                except TypeError:
                    # 如果aweme_info或statistics不是预期的字典类型，打印错误并跳过
                    print("TypeError: Expected a dictionary but got a different type.")
                except Exception as e:
                    # 捕获其他所有异常，打印错误信息并跳过
                    print(f"Unexpected error: {e}")
            offset += 10
            response.close()
            # txt相比于excel要快很多所以就用txt保存了
    print(ttwid)
    with open("results.txt", "a", encoding="utf-8") as file:
        file.write(f"{sum} {keyword}\n")


#async def main():
    #wb = load_workbook(filename='皖宁鲁冀藏苏_乡镇汇总.xlsx')
    #ws = wb.active
    #keywords = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    # 每5个关键字分成一组
    #group_size = 5
    #for i in range(0, len(keywords), group_size):
        #group = keywords[i:i+group_size]
        #tasks = [payapa(keyword) for keyword in group]
        #await asyncio.gather(*tasks)
        # await asyncio.sleep(1)
#asyncio.run(main())

wb = load_workbook(filename='皖宁鲁冀藏苏_乡镇汇总.xlsx')
ws = wb.active
keywords = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]

count = 0
proxies = get_proxies()
for keyword in keywords:
    if count == 8:
        proxies = get_proxies()
        payapa(keyword,proxies)
        count = 0
    else:
        payapa(keyword,proxies)
        count += 1





